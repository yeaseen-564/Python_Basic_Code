# import openpyxl as xl
# web = xl.load_workbook("transactions.xlsx")
# sheet = web["Sheet1"]
# cell = sheet["a1"]
# _cell = sheet.cell(1, 1)
# print(_cell.value)

# import openpyxl as xl
#
# web = xl.load_workbook("transactions.xlsx")
# sheet = web["Sheet1"]  # Here Sheet1 can not be in lower case
# cell = sheet["a1"]
# _cell = sheet.cell(1, 1)
# # for loop:
# for row in range(1, sheet.max_row + 1):
#     cell = sheet.cell(row, 3)
#     print(cell.value)

# import openpyxl
# wb = openpyxl.load_workbook("transactions.xlsx")
# # ws = wb.active
# work_sheet = wb["Sheet1"]
# cell = work_sheet["A1"]
#
# print(cell.value)
# print(wb.sheetnames)
#
# wb.save("transactions.xlsx")


import openpyxl

word_book = openpyxl.load_workbook("grade.xlsx")
work_sheet = word_book["Sheet1"]
cell_a1 = work_sheet["A1"]
print(cell_a1.value)
print(work_sheet)






